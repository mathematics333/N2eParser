import openpyxl

def copy_excel_data(source_file, target_file, sheet_name):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    
    # Load the target workbook
    target_workbook = openpyxl.load_workbook(target_file)
    
    # Get the sheet from the source workbook
    source_sheet = source_workbook[sheet_name]
    
    # Create a new sheet in the target workbook with the same name
    if sheet_name in target_workbook.sheetnames:
        target_sheet = target_workbook[sheet_name]
    else:
        target_sheet = target_workbook.create_sheet(sheet_name)
    
    # Iterate over the cells in the source sheet and copy the values to the target sheet
    for i, row in enumerate(source_sheet.iter_rows(values_only=False), start=1):
        for cell in row:
            if cell.column_letter == 'A':  # Check if the cell is in column A
                # Adjust the row index to start from row 4 in the target sheet
                target_sheet[f'A{i+3}'].value = cell.value
    
    # Save the target workbook
    target_workbook.save(target_file)
    print("Data copied successfully!")

# Specify the file paths and sheet name
source_file = 'NwOutput.xlsx'
target_file = 'acc-issues-import-template.xlsx'
sheet_name = 'Sheet1'

# Call the function to copy the data
copy_excel_data(source_file, target_file, sheet_name)
